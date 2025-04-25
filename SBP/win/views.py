
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .serializers import *
from rest_framework.authtoken.models import Token 
from datetime import datetime
from rest_framework.permissions import AllowAny, IsAuthenticated
from .models import *
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.generics import UpdateAPIView, ListAPIView, CreateAPIView
from rest_framework.pagination import PageNumberPagination
from django.db import IntegrityError
from django.shortcuts import get_object_or_404
from django.db import transaction
import logging
from django.db.models import Count
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from django.db.models import Exists, OuterRef
logger = logging.getLogger(__name__)

class UserApprovalView(APIView):
    """
    API для модерации пользователей (подтверждение/отклонение регистраций)
    Доступно только представителям ФСП (role.id=2)
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """
        Получение списка пользователей, ожидающих подтверждения
        Возвращает:
        - 403: если пользователь не администратор
        - 200: список пользователей в формате UserApprovalSerializer
        """
        # Проверка прав доступа (только для админов)
        if request.user.info.role.id != 2:
            return Response({'error': 'Недостаточно прав'}, status=status.HTTP_403_FORBIDDEN)
        
        # Оптимизированный запрос с выборкой связанных данных
        pending_users = UserInfo.objects.filter(
            role__id__in=[1, 2],  # Только роли 1 и 2 требуют подтверждения
            is_approved=False      # Только неподтвержденные пользователи
        ).select_related('user', 'role', 'region')  # Жадная загрузка
        
        serializer = UserApprovalSerializer(pending_users, many=True)
        return Response(serializer.data)
    
    def post(self, request):
        """
        Подтверждение/отклонение пользователя
        Параметры:
        - user_id: ID пользователя (обязательный)
        - action: "approve" или "reject" (обязательный)
        Возвращает:
        - 400: неверные параметры
        - 403: нет прав
        - 404: пользователь не найден
        - 200: успешное выполнение
        """
        # Проверка прав администратора
        if request.user.info.role.id != 2:
            return Response({'error': 'Недостаточно прав'}, status=status.HTTP_403_FORBIDDEN)
        
        user_id = request.data.get('user_id')
        action = request.data.get('action')  # Тип действия
        
        try:
            user_info = UserInfo.objects.get(user__id=user_id)
        except UserInfo.DoesNotExist:
            return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)
        
        # Логика обработки действий
        if action == 'approve':
            user_info.is_approved = True
            user_info.save()
            return Response({'message': 'Пользователь успешно подтвержден'})
        elif action == 'reject':
            # Внимание! Полное удаление пользователя
            user_info.user.delete()
            return Response({'message': 'Пользователь отклонен и удален'})
        else:
            return Response({'error': 'Неверное действие'}, status=status.HTTP_400_BAD_REQUEST)
        
class RegisterView(APIView):
    """
    API регистрации новых пользователей
    Особенности:
    - Для ролей 1 и 2 требуется подтверждение представителя ФСП
    - Для роли 0 сразу выдается токен
    """
    permission_classes = [AllowAny] 
    
    def post(self, request):
        """
        Создание нового пользователя
        Возвращает:
        - 400: ошибки валидации
        - 201: успешная регистрация
        """
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            
            # Базовая структура ответа
            response_data = {
                'user': {
                    'id': user.id,
                    'email': user.email,
                    'nickName': user.nickName,
                    'info': {
                        'name': user.info.name,
                        'surname': user.info.surname
                    }
                },
                'role': RoleSerializer(user.info.role).data
            }
            
            # Разная логика для разных ролей
            if user.info.role.id in [1, 2]:
                # Для ролей 1 и 2 - ожидание подтверждения
                response_data['message'] = 'Регистрация успешна. Ожидайте подтверждения администратором.'
                return Response(response_data, status=status.HTTP_201_CREATED)
            else:
                # Для роли 0 - сразу выдаем токен
                token, created = Token.objects.get_or_create(user=user)
                response_data['token'] = token.key
                response_data['message'] = 'Пользователь успешно зарегистрирован'
                return Response(response_data, status=status.HTTP_201_CREATED)
            
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class LoginView(APIView):
    """
    API для аутентификации пользователей и получения токена доступа.
    
    Особенности:
    - При успешной аутентификации возвращается токен, роль пользователя и базовая информация.
    - Доступ открыт для всех (не требует авторизации).
    """

    permission_classes = [AllowAny]

    def post(self, request):
        # Сериализуем и валидируем входные данные (логин, пароль)
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Получаем пользователя из валидированных данных
        user = serializer.validated_data['user']

        # Получаем или создаём токен для пользователя (DRF TokenAuthentication)
        token, created = Token.objects.get_or_create(user=user)

        # Получаем дополнительную информацию о пользователе
        user_info = user.info

        # Сериализуем роль пользователя для возврата в ответе
        role_serializer = RoleSerializer(user_info.role)

        # Формируем и возвращаем ответ с токеном, ролью и информацией о пользователе
        return Response({
            'token': token.key,
            'role': role_serializer.data,
            'user': {
                'id': user.id,
                'email': user.email,
                'nickName': user.nickName,
                'info': {
                    'name': user_info.name,
                    'surname': user_info.surname
                }
            }
        })
    
class CompetitionCreateView(APIView):
    
    """
    API endpoint для создания новых соревнований.
    
    Требуется аутентификация пользователя. При успешном создании автоматически
    создается запись организатора соревнования (CompetitionOrganizer) для текущего пользователя.
    
    Особенности обработки:
    - Поддерживает передачу названия дисциплины (автоматически конвертирует в ID)
    - Обрабатывает permissions в двух форматах (список объектов с id или список id)
    - Автоматически создает связанные CompetitionDate
    
    Возможные коды ответов:
    - 201: Соревнование успешно создано
    - 400: Ошибка валидации данных
    - 401: Пользователь не аутентифицирован
    """
    permission_classes = [IsAuthenticated]

    
    def post(self, request):
        """
        Обработка POST-запроса на создание соревнования.
        
        Параметры:
        - discipline: может быть строкой (название) или ID дисциплины
        - permissions: список регионов в формате [{'id': X}] или [X, Y, Z]
        - dates: объект с датами соревнования (start_date, end_date и др.)
        
        Автоматические действия:
        1. Конвертация названия дисциплины в ID (если передана строка)
        2. Нормализация формата permissions (в список ID)
        3. Создание записи CompetitionOrganizer для текущего пользователя
        4. Создание связанной записи CompetitionDate
        """
        # Преобразуем название дисциплины в ID если нужно
        if 'discipline' in request.data and isinstance(request.data['discipline'], str):
            try:
                discipline = Discipline.objects.get(name=request.data['discipline'])
                request.data['discipline'] = discipline.id
            except Discipline.DoesNotExist:
                return Response(
                    {"discipline": "Discipline not found"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Обрабатываем permissions - преобразуем в список ID
        if 'permissions' in request.data and isinstance(request.data['permissions'], list):
            try:
                # Извлекаем только ID из объектов регионов
                permissions_data = request.data['permissions']
                region_ids = [item['id'] for item in permissions_data if isinstance(item, dict) and 'id' in item]
                request.data['permissions'] = region_ids
            except (TypeError, KeyError):
                return Response(
                    {"permissions": "Invalid format - expected list of regions with ids"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        serializer = CompetitionSerializer(data=request.data)
        if serializer.is_valid():
            competition = serializer.save()
            
            # Получаем UserInfo текущего пользователя
            try:
                user_info = UserInfo.objects.get(user=request.user)
            except UserInfo.DoesNotExist:
                return Response(
                    {"error": "User profile not found"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Создаем запись организатора
            CompetitionOrganizer.objects.create(
                user=user_info,  # Передаем экземпляр UserInfo
                competition=competition,
                rated=False
            )
            
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class TeamCreateView(APIView):
    """
    API endpoint для создания новых команд в соревнованиях.
    
    Требуется аутентификация пользователя. При успешном создании возвращает
    основные данные о созданной команде.
    
    Особенности:
    - Проверяет заполненность профиля пользователя (UserInfo)
    - Обрабатывает ошибки создания с соответствующими HTTP-статусами
    - Возвращает структурированный ответ с ключевыми данными команды
    
    Возможные коды ответов:
    - 201: Команда успешно создана
    - 400: Ошибка валидации данных
    - 403: Профиль пользователя не заполнен
    - 500: Ошибка сервера при создании
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        POST-запрос для создания новой команды.
        
        Параметры запроса (в теле JSON):
        - competition: ID командного соревнования (обязательное)
        - name: Название команды (обязательное, макс. 100 символов)
        - description: Описание команды (необязательное)
        - is_private: Приватность команды (по умолчанию False)
        - captain_id: ID капитана (обязательно для модераторов)
        
        Возвращает при успехе:
        - team_id: ID созданной команды
        - name: Название команды
        - competition_id: ID соревнования
        - captain_id: ID капитана
        - max_members: Максимальное число участников
        """
        serializer = TeamCreateSerializer(
            data=request.data,
            context={'request': request}
        )
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            team = serializer.save()
            return Response({
                "team_id": team.id,
                "name": team.name,
                "competition_id": team.competition.id,
                "captain_id": team.captain.id,
                "max_members": team.max_members
            }, status=status.HTTP_201_CREATED)
        except UserInfo.DoesNotExist:
            return Response(
                {"error": "profile_incomplete", "detail": "Профиль пользователя не заполнен"},
                status=status.HTTP_403_FORBIDDEN
            )
        except Exception as e:
            return Response(
                {"error": "creation_error", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
            
class InvitationCreateView(APIView):
    """
    API endpoint для создания приглашений в команду.
    
    Доступ: 
    - Разрешен для всех аутентифицированных пользователей (AllowAny)
    - Но функционально работает только для капитанов команд
    
    Логика работы:
    - Проверяет права капитана на отправку приглашения
    - Создает приглашение со статусом "Ожидает"
    - Возвращает основные данные созданного приглашения
    
    Возможные коды ответов:
    - 201: Приглашение успешно создано
    - 400: Ошибка валидации данных
    - 403: Пользователь не является капитаном команды
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        POST-запрос для создания приглашения в команду.
        
        Параметры запроса (в теле JSON):
        - team_id: ID команды (обязательное)
        - user_id: ID приглашаемого пользователя (обязательное)
        
        Возвращает при успехе:
        - id: ID созданного приглашения
        - team_id: ID команды
        - user_id: ID приглашенного пользователя
        - status: Статус приглашения ("Ожидает")
        
        Особые проверки:
        - Только капитан команды может отправлять приглашения
        """
        serializer = InvitationCreateSerializer(
            data=request.data,
            context={'request': request}
        )
        
        if serializer.is_valid():
            # Проверяем что текущий пользователь - капитан команды
            team = serializer.validated_data['team']
            logger.debug(team.captain.id)
            logger.debug(request.user.id)
            if team.captain.id != request.user.id:
                return Response(
                    {"detail": "Только капитан команды может отправлять приглашения"},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            invitation = serializer.save()
            return Response({
                'id': invitation.id,
                'team_id': invitation.team.id,
                'user_id': invitation.user.id,
                'status': invitation.status
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class UserInvitationsView(APIView):
    """
    API endpoint для получения списка приглашений пользователя.
    
    Доступ:
    - Только для аутентифицированных пользователей
    
    Возвращает:
    - Список активных приглашений (со статусом "Ожидает")
    - Расширенные данные по каждому приглашению (команда, соревнование)
    
    Коды ответов:
    - 200: Успешный запрос, возвращает список приглашений
    - 401: Пользователь не аутентифицирован
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        GET-запрос для получения списка приглашений текущего пользователя.
        
        Фильтрация:
        - Только приглашения со статусом "Ожидает"
        - Только приглашения для текущего пользователя
        
        Возвращает:
        - Массив объектов приглашений с расширенной информацией:
          * Основные данные приглашения
          * Информация о команде
          * Название соревнования
          * Никнейм пользователя
        """
        invitations = Invitation.objects.filter(
            user=request.user.id,
            status='Ожидает'
        )
        serializer = InvitationSerializer(invitations, many=True)
        return Response(serializer.data)
    
class InvitationResponseView(UpdateAPIView):
    """
    API для ответа на приглашение в команду
    Доступные действия: accept (принять) или reject (отклонить)
    Только для владельца приглашения
    Только PATCH-запросы
    """
    serializer_class = InvitationResponseSerializer
    permission_classes = [IsAuthenticated]
    queryset = Invitation.objects.all()
    http_method_names = ['patch']  # Разрешаем только PATCH-метод

    def get_object(self):
        """
        Получает приглашение и проверяет права:
        - Приглашение должно существовать
        - Текущий пользователь должен быть получателем приглашения
        """
        try:
            invitation = super().get_object()
            # Проверка что пользователь отвечает на свое приглашение
            if invitation.user.id != self.request.user.id:
                raise PermissionDenied("Вы можете отвечать только на свои приглашения")
            return invitation
        except Invitation.DoesNotExist:
            raise status.HTTP_400_BAD_REQUEST

class RegionListView(ListAPIView):
    """
    API для получения списка регионов
    Доступ: без авторизации (AllowAny)
    Сортировка: по id (возрастание)
    """
    queryset = Region.objects.all().order_by('id')  # Оптимизированный запрос с сортировкой
    serializer_class = RegionSerializer
    permission_classes = [AllowAny]  # Доступно всем пользователям


class RoleListView(ListAPIView):
    """
    API для получения списка ролей пользователей
    Доступ: без авторизации (AllowAny) 
    Сортировка: по id (возрастание)
    """
    queryset = Role.objects.all().order_by('id')  # Оптимизированный запрос с сортировкой
    serializer_class = RoleSerializer
    permission_classes = [AllowAny]  # Доступно всем пользователям
    
class TeamApplicationCreateView(APIView):
    """
    API endpoint для создания заявки команды на участие в соревновании.
    
    Требуется аутентификация. Пользователь должен быть участником команды, 
    для которой создается заявка.
    
    Параметры запроса:
    - team_id (обязательный) - ID команды
    - competition (обязательный) - ID соревнования
    
    Возвращает:
    - 201 Created: при успешном создании заявки
    - 400 Bad Request: при ошибках валидации
    - 403 Forbidden: если пользователь не участник команды
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = TeamApplicationSerializer(
            data=request.data,
            context={'request': request}
        )
        
        if serializer.is_valid():
            application = serializer.save()
            return Response({
                'id': application.id,
                'team_id': application.team.id,
                'status': application.status,
                'reason': application.reason
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class TeamApplicationResponseView(UpdateAPIView):
    """
    API endpoint для обработки заявки команды (принятие/отклонение).
    
    Доступные действия:
    - accept - принять заявку (добавляет всех участников команды в соревнование)
    - reject - отклонить заявку (требуется указать причину)
    
    Требуется аутентификация. Доступно только для заявок со статусом 'pending'.
    
    Метод: PATCH
    
    Параметры:
    - action (обязательный) - действие (accept/reject)
    - reason (обязательный для reject) - причина отклонения
    
    Возвращает:
    - 200 OK: при успешном обновлении
    - 400 Bad Request: при ошибках валидации
    - 403 Forbidden: если заявка уже обработана
    """
    serializer_class = TeamApplicationResponseSerializer
    permission_classes = [IsAuthenticated]
    queryset = TeamApplication.objects.all()
    http_method_names = ['patch']

    def perform_update(self, serializer):   
        serializer.save()
        

class CompetitionListView(ListAPIView):
    """
    API endpoint для получения списка доступных соревнований.
    
    Доступно без аутентификации. Не включает соревнования со статусом 'pending'.
    
    Возвращает список соревнований с детальной информацией:
    - ID, название, описание, статус
    - Дисциплина
    - Даты проведения
    """
    permission_classes = [AllowAny]
    serializer_class = CompetitionSerializer
    queryset = Competition.objects.exclude(status='pending').select_related(
        'discipline',
        'dates'
    )
    
class StandardResultsSetPagination(PageNumberPagination):
    """Стандартная пагинация с размером страницы 10 (макс. 100)"""
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class FAQListView(ListAPIView):
    """
    Список часто задаваемых вопросов (FAQ)
    Доступ: без авторизации
    Пагинация: 10 на страницу (настраивается через page_size)
    """
    permission_classes = [AllowAny]
    queryset = FAQ.objects.all()
    serializer_class = FAQSerializer
    pagination_class = StandardResultsSetPagination


class NewsPagination(PageNumberPagination):
    """Пагинация новостей: 10 на страницу (макс. 100)"""
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class NewsListView(ListAPIView):
    """
    Список новостей с сортировкой по дате (новые сначала)
    Доступ: без авторизации
    Поиск: по title и content
    Пагинация: 10 на страницу
    """
    permission_classes = [AllowAny]
    queryset = News.objects.all().order_by('-created_at')
    serializer_class = NewsSerializer
    pagination_class = NewsPagination
    search_fields = ['title', 'content']
    
class UserApplicationCreateView(CreateAPIView):
    """
    Создание заявки пользователя на индивидуальное соревнование.
    Доступ: только для аутентифицированных пользователей.
    
    Ошибки:
    - 400: Если заявка уже существует или данные невалидны
    - 403: Если пользователь не соответствует требованиям соревнования
    """
    serializer_class = UserApplicationSerializer
    permission_classes = [IsAuthenticated]
    queryset = UserApplication.objects.all()

    def perform_create(self, serializer):
        """Обработка создания заявки с проверкой уникальности"""
        try:
            serializer.save()
        except IntegrityError:
            raise ValidationError("Вы уже подавали заявку на это соревнование")


class DisciplineListView(ListAPIView):
    """
    Получение списка всех дисциплин.
    Доступ: без авторизации.
    Сортировка: по ID.
    """
    queryset = Discipline.objects.all().order_by('id')
    serializer_class = DisciplineSerializer
    permission_classes = [AllowAny]
    
class ApplicationDecisionView(UpdateAPIView):
    """
    API для принятия решения по заявке пользователя на соревнование.
    
    Доступ:
    - Только для аутентифицированных организаторов данного соревнования
    
    Параметры:
    - action (обязательный): accept/reject - решение по заявке
    - reason (опционально): причина отказа (обязателен при reject)
    
    Ошибки:
    - 403: Если пользователь не организатор
    - 400: При превышении лимита участников или невалидных данных
    """
    serializer_class = ApplicationDecisionSerializer
    permission_classes = [IsAuthenticated]
    queryset = UserApplication.objects.all()

    def get_object(self):
        """Получает заявку и проверяет права организатора"""
        application = get_object_or_404(UserApplication, pk=self.kwargs['pk'])
        user_info = self.request.user
        
        # Проверка прав организатора
        if not CompetitionOrganizer.objects.filter(
            user=user_info.id,
            competition=application.competition
        ).exists():
            raise PermissionDenied("Вы не являетесь организатором этого соревнования")
        
        return application

    def perform_update(self, serializer):
        """Обрабатывает решение по заявке"""
        application = self.get_object()
        action = serializer.validated_data['action']
        reason = serializer.validated_data.get('reason', '')

        if action == 'accept':
            # Проверка лимита участников
            if application.competition.participants.count() >= application.competition.max_participants:
                raise ValidationError("Достигнуто максимальное количество участников")
            
            # Принятие заявки
            application.status = 'accepted'
            application.reason = None
            # Добавление участника
            CompetitionParticipant.objects.get_or_create(
                competition=application.competition,
                participant=application.user
            )
        else:
            # Отклонение заявки
            application.status = 'rejected'
            application.reason = reason
        
        application.save()
        
class OrganizerUserApplicationsListView(ListAPIView):
    """
    API для получения списка заявок на соревнования, где пользователь является организатором.
    
    Возвращает:
    - Список заявок со статусом 'pending' для соревнований, где пользователь организатор
    - Информацию о пользователях и соревнованиях
    
    Доступ:
    - Только для аутентифицированных организаторов
    
    Фильтрация:
    - Только ожидающие рассмотрения заявки (status='pending')
    """
    serializer_class = UserApplicationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Возвращает заявки на соревнования, где пользователь организатор"""
        user_info = self.request.user
        # Получаем ID соревнований пользователя как организатора
        organized_competitions = CompetitionOrganizer.objects.filter(
            user=user_info.id
        ).values_list('competition', flat=True)
        
        return UserApplication.objects.filter(
            competition__in=organized_competitions,
            status='pending'
        ).select_related('user', 'competition')  # Оптимизация запросов

class UserListView(ListAPIView):
    """
    API для получения списка обычных пользователей (role_id=0)
    
    Возвращает:
    - Список пользователей с основной информацией
    - Сортировка по ID
    
    Доступ:
    - Без авторизации
    
    Фильтрация:
    - Только пользователи с role_id=0 (обычные пользователи)
    """
    queryset = UserInfo.objects.filter(role_id=0).order_by('-rating')
    serializer_class = UserInfoSerializer
    permission_classes = [AllowAny]
    

class PublicTeamsView(APIView):
    """
    API для получения списка публичных команд
    
    Возвращает:
    - Полную информацию о публичных командах (is_private=False)
    - Данные о соревновании и датах проведения
    - Информацию о капитане команды
    - Количество участников
    
    Поля ответа:
    - id: ID команды
    - name: Название команды
    - description: Описание команды
    - competition: Данные соревнования (id, name, dates)
    - captain: Данные капитана (id, nickName)
    - max_members: Максимальное количество участников
    - current_members: Текущее количество участников
    - is_private: Флаг приватности (всегда False)
    
    Оптимизации:
    - select_related для связей competition, captain, dates
    - prefetch_related для members
    - annotate для подсчета участников
    
    Доступ:
    - Без авторизации
    """
    permission_classes = [AllowAny]
    
    def get(self, request):
        teams = Team.objects.filter(is_private=False).select_related(
            'competition', 
            'captain',
            'captain__user',  # Добавляем связь с пользователем капитана
            'competition__dates'  # Добавляем связь с датами соревнования
        ).prefetch_related(
            'members'
        ).annotate(
            members_count=Count('members')
        )
        
        data = []
        for team in teams:
            team_data = {
                'id': team.id,
                'name': team.name,
                'description': team.description,
                'competition': {
                    'id': team.competition.id,
                    'name': team.competition.name,
                    'dates': {
                        'start_date': team.competition.dates.start_date if hasattr(team.competition, 'dates') else None,
                        'end_date': team.competition.dates.end_date if hasattr(team.competition, 'dates') else None,
                        'registration_start': team.competition.dates.registration_start if hasattr(team.competition, 'dates') else None,
                        'registration_end': team.competition.dates.registration_end if hasattr(team.competition, 'dates') else None,
                    }
                },
                'captain': {
                    'id': team.captain.id if team.captain else None,
                    'nickName': team.captain.user.nickName if team.captain and hasattr(team.captain, 'user') else None
                },
                'max_members': team.max_members,
                'current_members': team.members_count,  # Используем аннотированное значение
                'is_private': team.is_private,
                'members_count': team.members_count  # Дублируем для совместимости
            }
            data.append(team_data)
        
        return Response({
            'count': len(data),
            'teams': data
        }, status=status.HTTP_200_OK)
        
class CaptainVacancyResponsesView(APIView):
    """
    API для капитанов команд по управлению откликами на вакансии
    
    Возвращает:
    - Список ожидающих (pending) откликов на вакансии в командах, где пользователь является капитаном
    
    Поля ответа:
    - count: общее количество откликов
    - responses: список откликов (сериализованные данные VacancyResponse)
    
    Ошибки:
    - 403: если пользователь не является капитаном ни одной команды
    
    Оптимизации:
    - select_related для связи с командой
    - фильтрация только pending-откликов
    
    Доступ:
    - Только для аутентифицированных капитанов команд
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Проверяем, является ли пользователь капитаном какой-либо команды
        user_teams = Team.objects.filter(captain=request.user.id)
        
        if not user_teams.exists():
            return Response(
                {"error": "not_a_captain", "detail": "Вы не являетесь капитаном ни одной команды"},
                status=status.HTTP_403_FORBIDDEN
            )

        # Получаем все отклики для команд пользователя-капитана
        responses = VacancyResponse.objects.filter(
            team__in=user_teams.values_list('id', flat=True),
            status = 'pending'
        ).select_related('team')  # Оптимизация запросов к БД

        serializer = VacancyResponseSerializer(responses, many=True)
        
        return Response({
            'count': responses.count(),
            'responses': serializer.data
        }, status=status.HTTP_200_OK)
        
class ResponseToPublicView(APIView):
    """
    API для отправки отклика на публичную вакансию в команде
    
    Позволяет пользователю:
    - Отправить отклик на участие в команде
    - Автоматически привязывает отклик к профилю пользователя
    
    Параметры запроса:
    - team_id (обязательный): ID команды
    - message (опциональный): сопроводительное сообщение
    
    Возвращает:
    - 201: при успешном создании отклика (возвращает данные отклика)
    - 400: при ошибках валидации или отсутствии профиля
    - 401: для неаутентифицированных пользователей

    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            user_info = UserInfo.objects.get(user=request.user)
        except UserInfo.DoesNotExist:
            return Response(
                {"detail": "Профиль пользователя не найден"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Добавляем user_info в данные для сериализатора
        request.data['user'] = user_info.id
        
        serializer = VacancyResponseSerializer(
            data=request.data,
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ResponseActionView(APIView):
    """
    API для обработки откликов на вакансии в команде (принять/отклонить)
    
    Доступ:
    - Только для аутентифицированных капитанов команды
    
    Параметры запроса:
    - response_id (обязательный): ID отклика на вакансию
    - action (обязательный): accept/reject - действие по отклику
    
    Возможные ответы:
    - 200: Успешное выполнение действия
    - 400: Неверные данные или нет мест в команде
    - 403: Пользователь не является капитаном
    - 404: Отклик не найден
    
    Логика работы:
    - При accept: добавляет пользователя в команду (если есть места)
    - При reject: отклоняет заявку без дополнительных действий
    - Ведет логирование для отладки
    
    Валидации:
    - Проверка прав капитана
    - Проверка наличия мест в команде
    - Проверка существования отклика
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        serializer = ResponseActionSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        response = get_object_or_404(VacancyResponse, id=serializer.validated_data['response_id'])
        team = response.team
        logger.debug(f'response : {response}')
        logger.debug(f"team : {team}")
        logger.debug(f'team captain : {team.captain.id}')
        logger.debug(f'request.user.id : {request.user.id}')
        # Проверяем, что текущий пользователь - капитан команды
        if request.user.id != team.captain.id:
            return Response(
                {"detail": "Только капитан команды может обрабатывать заявки"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        action = serializer.validated_data['action']
        
        if action == 'accept':
            # Проверяем, есть ли место в команде
            if team.current_members >= team.max_members:
                return Response(
                    {"detail": "В команде нет свободных мест"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Добавляем пользователя в команду
            team.members.add(response.user)
            team.current_members = team.members.count()
            team.save()
            
            # Обновляем статус заявки
            response.status = VacancyResponse.ACCEPTED
            response.save()
            
            return Response(
                {"detail": "Заявка принята, пользователь добавлен в команду"},
                status=status.HTTP_200_OK
            )
        
        elif action == 'reject':
            response.status = VacancyResponse.REJECTED
            response.save()
            
            return Response(
                {"detail": "Заявка отклонена"},
                status=status.HTTP_200_OK
            )
            
class UserProfileView(APIView):
    """
    API для работы с профилем пользователя
    
    GET:
    - Возвращает полную информацию о пользователе и его профиле
    - Включает данные:
      * Основные данные пользователя (email, nickName)
      * Информацию профиля (ФИО, регион, роль, дата рождения)
      * Названия региона и роли (в дополнение к ID)
    
    PATCH:
    - Частичное обновление данных пользователя и профиля
    - Принимает данные в формате:
      {
        "user": {"email": "...", "nickName": "..."},
        "info": {"surname": "...", "region": id, ...}
      }
    - Обновляет только переданные поля
    
    Доступ:
    - Только для аутентифицированных пользователей
    - Каждый пользователь может работать только со своим профилем
    """
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = request.user  # Получаем объект пользователя, а не только ID
        user_info = get_object_or_404(UserInfo, user=user)
        
        # Сериализуем данные пользователя
        user_serializer = UserUpdateSerializer(user)  # Теперь передаем объект пользователя
        
        # Сериализуем данные профиля с регионом
        info_serializer = UserInfoUpdateSerializer(user_info)
        
        # Получаем данные региона
        region = user_info.region
        region_serializer = RegionSerializer(region) if region else None
        
        # Получаем данные роли
        role = user_info.role
        role_serializer = RoleSerializer(role) if role else None
        
        response_data = {
            'user': user_serializer.data,
            'info': info_serializer.data
        }
        
        # Добавляем название региона в ответ
        if region_serializer:
            response_data['info']['region_name'] = region_serializer.data['name']
        
        # Добавляем название роли в ответ
        if role_serializer:
            response_data['info']['role_name'] = role_serializer.data['name']
        
        return Response(response_data)
    
    def patch(self, request):
        # Получаем текущего пользователя и его профиль
        user = request.user
        user_info = get_object_or_404(UserInfo, user=user.id)
        
        # Сериализуем данные
        serializer = UserProfileUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # Обновляем данные User
        user_data = serializer.validated_data.get('user', {})
        if user_data:
            user_serializer = UserUpdateSerializer(user, data=user_data, partial=True)
            if user_serializer.is_valid():
                user_serializer.save()
            else:
                return Response(user_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # Обновляем данные UserInfo
        info_data = serializer.validated_data.get('info', {})
        if info_data:
            info_serializer = UserInfoUpdateSerializer(user_info, data=info_data, partial=True)
            if info_serializer.is_valid():
                info_serializer.save()
            else:
                return Response(info_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(
            {"detail": "Данные успешно обновлены"},
            status=status.HTTP_200_OK
        )
        
class ParticipationHistoryView(APIView):
    """
    API для получения истории участия пользователя в соревнованиях
    
    Возвращает:
    - Статистику участий (общее количество, победы, подиумы, текущий рейтинг)
    - Подробную историю участия с результатами
    
    Поля ответа:
    - stats: {
        total_participations: общее количество участий
        wins: количество побед (1 место)
        podiums: количество попаданий в топ-3
        current_rating: текущий рейтинг пользователя
      }
    - history: список участий {
        competition: краткая информация о соревновании
        result: занятое место
      }
    
    Доступ:
    - Только для аутентифицированных пользователей
    
    Ошибки:
    - 404: если профиль пользователя не найден
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            user_info = UserInfo.objects.get(user=request.user.id)
        except UserInfo.DoesNotExist:
            return Response(
                {"detail": "Профиль пользователя не найден"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        participations = CompetitionParticipant.objects.filter(
            participant=user_info
        ).select_related('competition').annotate(
            total_participants=Count('competition__participants')
        )
        
        stats = {
            'total_participations': participations.count(),
            'wins': participations.filter(result=1).count(),
            'podiums': participations.filter(result__lte=3).count(),
            'current_rating': user_info.rating,
        }
        
        serializer = ParticipationHistorySerializer(participations, many=True)
        
        return Response({
            'stats': stats,
            'history': serializer.data
        })
        
class OrganizedCompetitionsView(APIView):
    """
    API для получения списка соревнований, где пользователь является организатором
    
    Возвращает:
    - Список соревнований с основной информацией:
      * ID, название, дисциплина
      * Тип (индивидуальное/командное)
      * Статус соревнования
      * Флаг rated (оценено ли соревнование)
    
    Ошибки:
    - 404: если профиль пользователя не найден
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Получаем UserInfo текущего пользователя
        user_info = get_object_or_404(UserInfo, user=request.user)
        
        # Получаем все соревнования где пользователь организатор
        organizers = CompetitionOrganizer.objects.filter(
            user=user_info
        ).select_related('competition', 'competition__discipline')
        
        serializer = OrganizerCompetitionSerializer(organizers, many=True)
        
        return Response({
            'competitions': serializer.data
        })
        
class DistributeResultsView(APIView):
    """
    API для распределения результатов завершенного соревнования
    
    Требования:
    - Пользователь должен быть организатором соревнования
    - Соревнование должно быть в статусе 'completed'
    
    Параметры запроса:
    - competition_id: ID соревнования
    - results: массив объектов с user_id и result (занятое место)
    
    Логика работы:
    1. Проверяет валидность данных
    2. Проверяет права доступа (организатор)
    3. Проверяет статус соревнования
    4. Обновляет результаты участников в транзакции
    5. Помечает соревнование как оцененное
    
    Возвращает:
    - 200: при успешном обновлении
    - 400: при ошибках валидации или неверном статусе
    - 403: если пользователь не организатор
    - 404: если соревнование не найдено
    """
    permission_classes = [IsAuthenticated]
    
    @transaction.atomic
    def post(self, request):
        serializer = CompetitionResultsSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        competition_id = serializer.validated_data['competition_id']
        results_data = serializer.validated_data['results']
        
        # Получаем соревнование
        competition = get_object_or_404(Competition, id=competition_id)
        
        # Проверяем что соревнование завершено
        if competition.status != 'completed':
            return Response(
                {"detail": "Нельзя распределить места для незавершенного соревнования"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Проверяем что пользователь организатор этого соревнования
        user_info = get_object_or_404(UserInfo, user=request.user)
        is_organizer = CompetitionOrganizer.objects.filter(
            user=user_info,
            competition=competition
        ).exists()
        
        if not is_organizer:
            return Response(
                {"detail": "Только организатор может распределять места"},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Обновляем результаты участников
        for result_data in results_data:
            participant = get_object_or_404(
                CompetitionParticipant,
                competition=competition,
                participant_id=result_data['user']
            )
            participant.result = result_data['result']
            participant.save()
        
        # Помечаем что организатор оценил соревнование
        organizer = CompetitionOrganizer.objects.get(
            user=user_info,
            competition=competition
        )
        organizer.rated = True
        organizer.save()
        
        return Response(
            {"detail": "Места успешно распределены"},
            status=status.HTTP_200_OK
        )
        
class UserTeamsView(APIView):
    """
    API для получения списка команд пользователя
    
    Возвращает:
    - Количество команд пользователя
    - Подробную информацию о каждой команде
    
    Особенности:
    - Добавляет аннотацию is_register для проверки регистрации
    
    Доступ:
    - Только для аутентифицированных пользователей
    
    Ошибки:
    - 404: если профиль пользователя не найден
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user_info = get_object_or_404(UserInfo, user_id=request.user.id)
        
        teams = Team.objects.filter(
            members=user_info.user_id
        ).select_related(
            'competition',
            'competition__discipline'
        ).prefetch_related(
            'members',
            'members__user'
        ).annotate(
            is_register=Exists(TeamApplication.objects.filter(team=OuterRef('pk'))))
        
        logger.debug(f"User ID: {request.user.id}, UserInfo ID: {user_info.id}")
        logger.debug(f"Teams found: {teams.count()}")
        
        serializer = TeamListSerializer(teams, many=True)
        return Response({
            'count': teams.count(),
            'teams': serializer.data
        })
        
class PendingCompetitionsView(APIView):
    """
    API для получения списка соревнований, ожидающих подтверждения (со статусом 'pending')
    
    Возвращает:
    - Список соревнований с полной информацией:
      * Основные данные (ID, название, описание)
      * Информацию о дисциплине
      * Даты проведения
    
    Особенности:
    - Использует оптимизированные запросы (select_related, prefetch_related)
    - Возвращает только соревнования в статусе 'pending'
    
    Доступ:
    - Только для аутентифицированных пользователей
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Получаем все соревнования со статусом 'pending'
        competitions = Competition.objects.filter(status='pending').select_related(
            'discipline'
        ).prefetch_related('dates')
        
        serializer = CompetitionSerializer(competitions, many=True)
        
        return Response({
            'competitions': serializer.data
        })
        
class OrganizerTeamApplicationsListView(ListAPIView):
    """
    API для получения списка заявок команд на соревнования, где пользователь является организатором
    
    Возвращает:
    - Список заявок команд с полной информацией:
      * Данные о команде
      * Информацию о соревновании
      * Статус заявки
    
    Особенности:
    - Фильтрует только заявки со статусом 'pending'
    - Использует оптимизированные запросы (select_related)
    - Возвращает только заявки на соревнования, где пользователь организатор
    
    Доступ:
    - Только для аутентифицированных организаторов
    """
    serializer_class = TeamApplicationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user_info = self.request.user
        # Получаем соревнования, где пользователь организатор
        organized_competitions = CompetitionOrganizer.objects.filter(
            user=user_info.id
        ).values_list('competition', flat=True)
        
        return TeamApplication.objects.filter(
            competition__in=organized_competitions,
            status='pending'  # или 'pending' в зависимости от вашей логики
        ).select_related('team', 'competition')
        
class CompetitionDecisionView(APIView):
    """
    API для подтверждения или отклонения соревнований модератором
    
    Требования:
    - Пользователь должен иметь роль модератора (role.id = 2)
    - Соревнование должно быть в статусе 'pending'
    
    Параметры запроса:
    - competition_id: ID соревнования
    - action: 'accept' (подтвердить) или 'reject' (отклонить)
    
    Логика работы:
    - При подтверждении (accept):
      * Меняет статус соревнования на 'upcoming'
    - При отклонении (reject):
      * Удаляет все связанные записи организаторов
      * Удаляет само соревнование
    
    Возвращает:
    - 200: при успешном выполнении действия
    - 400: при ошибках валидации
    - 403: если пользователь не модератор
    - 404: если соревнование или профиль не найдены
    
    Особенности:
    - Использует атомарную транзакцию для безопасности данных
    - Жёстко устанавливает статус 'pending' для входящих данных
    """
    permission_classes = [IsAuthenticated]
    
    
    @transaction.atomic
    def post(self, request):
        request.data['status'] = 'pending'  # Жёстко перезаписываем
        serializer = CompetitionDecisionSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        competition_id = serializer.validated_data['competition_id']
        action = serializer.validated_data['action']
        
        # Получаем соревнование
        competition = get_object_or_404(Competition, id=competition_id, status='pending')
        
        # Проверяем что пользователь имеет право подтверждать соревнования
        try:
            user_info = UserInfo.objects.get(user=request.user)
            if user_info.role.id != 2:  # Проверка что пользователь модератор
                return Response(
                    {"detail": "Только представители ФСП могут подтверждать соревнования"},
                    status=status.HTTP_403_FORBIDDEN
                )
        except UserInfo.DoesNotExist:
            return Response(
                {"detail": "Профиль пользователя не найден"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        if action == 'accept':
            # Обновляем статус соревнования
            competition.status = 'upcoming'
            competition.save()
            
            return Response(
                {"detail": "Соревнование подтверждено", "competition_id": competition.id},
                status=status.HTTP_200_OK
            )
        
        elif action == 'reject':
            # Удаляем записи организаторов
            CompetitionOrganizer.objects.filter(competition=competition).delete()
            
            # Удаляем само соревнование
            competition.delete()
            
            return Response(
                {"detail": "Соревнование отклонено и удалено"},
                status=status.HTTP_200_OK
            )
            
class CompetitionParticipantsView(APIView):
    """
    API для получения списка участников соревнования
    
    Возвращает:
    - Основную информацию о соревновании (ID, название, тип)
    - Список участников:
      * Для индивидуальных соревнований - список пользователей
      * Для командных соревнований - список команд с участниками
    
    Особенности:
    - Разные форматы данных для индивидуальных и командных соревнований
    - Использует оптимизированные запросы (select_related, prefetch_related)
    - Возвращает только подтвержденных участников (status='approved')
    
    Доступ:
    - Только для аутентифицированных пользователей
    
    Ошибки:
    - 404: если соревнование не найдено
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, competition_id):
        competition = get_object_or_404(Competition, id=competition_id)
        
        if competition.type == 'individual':
            # Для индивидуальных соревнований
            participants = UserApplication.objects.filter(
                competition=competition,
                status='approved'
            ).select_related('user__user')
            
            serializer = IndividualParticipantSerializer(participants, many=True)
            
        else:
            # Для командных соревнований
            participants = TeamApplication.objects.filter(
                competition=competition,
                status='approved'
            ).select_related('team__captain').prefetch_related('team__members__user')
            
            serializer = TeamParticipantSerializer(participants, many=True)
        
        return Response({
            'competition_id': competition.id,
            'competition_name': competition.name,
            'competition_type': competition.type,
            'participants': serializer.data
        })
        
class RegionalRepresentativesView(ListAPIView):
    """
    API для получения списка региональных представителей
    
    Возвращает:
    - Список подтвержденных пользователей с ролью регионального представителя (role_id=1)
    - Основную информацию о представителях:
      * ФИО
      * Email
      * Название региона
    
    Особенности:
    - Использует оптимизированные запросы (select_related)
    - Фильтрует только подтвержденных представителей (is_approved=True)
    
    Доступ:
    - Без авторизации (можно изменить на IsAuthenticated при необходимости)
    """
    permission_classes = [AllowAny]  # Или [IsAuthenticated] если нужно ограничить доступ
    serializer_class = RegionalRepresentativeSerializer
    
    def get_queryset(self):
        return UserInfo.objects.filter(
            role_id=1,  # Фильтр по role=1 (региональные представители)
            is_approved=True  # Только подтвержденные пользователи
        ).select_related('user', 'region')  # Оптимизация запросов
        
class CompetitionStatusView(APIView):
    """
    API для обновления статусов соревнований на основе текущего времени
    
    Принимает:
    - Временную метку в ISO 8601 формате (например, "2025-03-02T21:00:00Z")
    
    Возвращает:
    - Информацию о каждом соревновании:
      * ID и название
      * Старый и новый статус
      * Флаг изменения статуса
    - Общее количество обновленных соревнований
    - Временные метки клиента и сервера
    
    Логика работы:
    1. Проверяет корректность переданного времени
    2. Для каждого соревнования определяет актуальный статус:
       - registration: если текущее время в периоде регистрации
       - running: если время проведения соревнования
       - finished: если время окончания прошло
       - waiting: если до начала регистрации
    3. Обновляет статусы при необходимости
    
    Особенности:
    - Не изменяет статус 'pending' (ожидающие подтверждения)
    - Использует оптимизированные запросы (select_related, only)
    - Поддерживает временные зоны (UTC)
    
    Доступ:
    - Без авторизации
    """
    permission_classes = [AllowAny]
    
    def post(self, request):
        # Получаем время из запроса
        client_time_str = request.data.get('time')
        logger.debug(f"Received time: {client_time_str}")
        
        if not client_time_str:
            return Response(
                {"error": "Параметр 'time' обязателен в формате ISO 8601 (например, 2025-03-02T21:00:00Z)"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Парсим время напрямую (datetime.fromisoformat в Python 3.11+ поддерживает 'Z')
            client_time = datetime.fromisoformat(client_time_str)
            # Если время наивное (без часового пояса), добавляем UTC
            if client_time.tzinfo is None:
                client_time = timezone.make_aware(client_time, timezone.utc)
        except ValueError:
            return Response(
                {"error": "Неверный формат времени. Используйте ISO 8601 (например, 2025-03-02T21:00:00Z)"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Получаем все соревнования с датами
        competitions = Competition.objects.filter(
            dates__isnull=False
        ).select_related('dates').only(
            'id', 'status', 'name', 'dates__start_date', 
            'dates__end_date', 'dates__registration_start',
            'dates__registration_end'
        )
        
        updated_competitions = []
        
        for comp in competitions:
            original_status = comp.status
            new_status = original_status
            
            # Если текущий статус pending - не меняем
            if original_status == 'pending':
                updated_competitions.append({
                    'id': comp.id,
                    'name': comp.name,
                    'original_status': original_status,
                    'new_status': new_status,
                    'status_changed': False
                })
                continue
            
            dates = comp.dates
            
            # Определяем новый статус
            if dates.registration_start <= client_time <= dates.registration_end:
                new_status = 'registration'
            elif dates.start_date <= client_time <= dates.end_date:
                new_status = 'running'
            elif client_time > dates.end_date:
                new_status = 'finished'
            else:
                new_status = 'waiting'
            
            # Обновляем если статус изменился
            status_changed = new_status != original_status
            if status_changed:
                comp.status = new_status
                comp.save(update_fields=['status'])
            
            updated_competitions.append({
                'id': comp.id,
                'name': comp.name,
                'original_status': original_status,
                'new_status': new_status,
                'status_changed': status_changed
            })
        
        return Response({
            'client_time': client_time_str,
            'server_time': timezone.now().isoformat(),
            'competitions_updated': len([c for c in updated_competitions if c['status_changed']]),
            'competitions': updated_competitions
        })
        
class UserVacancyResponsesView(APIView):
    """
    API для получения списка откликов пользователя на вакансии в командах
    
    Возвращает:
    - Общее количество откликов
    - Список откликов с детальной информацией:
      * Данные о команде и соревновании
      * Текст отклика
      * Статус и его текстовое представление
    
    Особенности:
    - Использует оптимизированные запросы (select_related)
    - Возвращает только отклики текущего пользователя
    - Включает человекочитаемые названия статусов
    
    Доступ:
    - Только для аутентифицированных пользователей
    
    Ошибки:
    - 404: если профиль пользователя не найден
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            # Получаем UserInfo текущего пользователя
            user_info = request.user.info
        except AttributeError:
            return Response(
                {"detail": "Профиль пользователя не найден"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Получаем все отклики пользователя с предварительной загрузкой связанных данных
        responses = VacancyResponse.objects.filter(
            user=user_info
        ).select_related(
            'team',
            'team__competition'
        )
        
        serializer = UserVacancyResponseSerializer(responses, many=True)
        
        return Response({
            'count': responses.count(),
            'responses': serializer.data
        })
        
class RegionCompetitionsView(APIView):
    """
    API для получения активных соревнований в указанном регионе
    
    Параметры запроса (POST):
    - region: название региона (строка)
    
    Возвращает:
    - ID найденного региона
    - Название региона
    - Количество доступных соревнований
    - Список соревнований с краткой информацией:
      * ID, название, дисциплина
      * Тип и статус соревнования
    
    Особенности:
    - Принимает название региона в теле запроса
    - Находит соответствующий регион в базе данных
    - Фильтрует соревнования по наличию региона в permissions
    - Исключает соревнования в статусах 'pending' и 'finished'
    - Использует оптимизированные запросы (select_related)
    
    Доступ:
    - Без авторизации
    
    Ошибки:
    - 400: если параметр region не указан или регион не найден
    """
    permission_classes = [AllowAny]
    
    def post(self, request):
        region_name = request.data.get('region')
        
        if not region_name:
            return Response(
                {"error": "Параметр 'region' обязателен в теле запроса"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Ищем регион по названию (регистронезависимо)
            region = Region.objects.get(name__iexact=region_name)
        except Region.DoesNotExist:
            return Response(
                {"error": f"Регион '{region_name}' не найден"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Получаем активные соревнования для региона
        competitions = Competition.objects.filter(
            permissions__contains=[region.id],
        ).exclude(
            status__in=['pending', 'finished']
        ).select_related('discipline')
        
        serializer = CompetitionShortSerializer(competitions, many=True)
        
        return Response({
            'region_id': region.id,
            'region_name': region.name,
            'count': competitions.count(),
            'competitions': serializer.data
        })    

class CompetitionParticipantsStructuredExportAPI(APIView):
    """
    API для экспорта структурированных данных о соревнованиях и участниках в Excel
    
    Параметры:
    - competition_id (опциональный): ID конкретного соревнования
    
    Возвращает:
    - Excel файл с детализированной информацией:
      * Основные данные соревнования
      * Список индивидуальных участников с результатами
      * Состав команд с информацией об участниках
    
    Особенности:
    - Поддерживает экспорт как одного соревнования, так и всех соревнований
    - Форматирует данные с заголовками, стилями и границами
    - Автоматически настраивает ширину столбцов
    - Оптимизирует запросы к базе данных (select_related, prefetch_related)
    - Группирует участников по типу участия (индивидуальные/командные)
    
    Доступ:
    - Без авторизации
    
    Ошибки:
    - 404: если соревнование не найдено
    - 500: при внутренних ошибках сервера
    """
    permission_classes = [AllowAny]

    def get(self, request, competition_id=None):
        try:
            # Получаем данные из базы
            if competition_id:
                competitions = Competition.objects.filter(pk=competition_id)
            else:
                competitions = Competition.objects.all().order_by('id')
            
            if not competitions.exists():
                return Response(
                    {"error": "Соревнования не найдены"},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Создаем Excel файл в памяти
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Соревнования и участники"

            # Стили для оформления
            header_font = Font(bold=True, size=12)
            comp_header_font = Font(bold=True, size=12, color='003366')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')

            # Начальная строка
            current_row = 1

            for comp in competitions:
                # Заголовок соревнования
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
                comp_title = f"Соревнование: {comp.name} (ID: {comp.id})"
                ws.cell(row=current_row, column=1, value=comp_title).font = comp_header_font
                current_row += 1

                # Основная информация о соревновании
                comp_data = [
                    ["Тип", comp.get_type_display()],
                    ["Формат", comp.get_competition_type_display()],
                    ["Дисциплина", comp.discipline.name],
                    ["Статус", comp.status],
                    ["Макс. участников", comp.max_participants],
                    ["Описание", comp.description[:200] + "..." if len(comp.description) > 200 else comp.description]
                ]

                # Добавляем даты, если они есть
                try:
                    comp_dates = CompetitionDate.objects.get(competition=comp)
                    comp_data.extend([
                        ["Даты проведения", f"{comp_dates.start_date.strftime('%d.%m.%Y %H:%M')} - {comp_dates.end_date.strftime('%d.%m.%Y %H:%M')}"],
                        ["Даты регистрации", f"{comp_dates.registration_start.strftime('%d.%m.%Y %H:%M')} - {comp_dates.registration_end.strftime('%d.%m.%Y %H:%M')}"]
                    ])
                except CompetitionDate.DoesNotExist:
                    pass

                # Записываем данные о соревновании
                for row_data in comp_data:
                    ws.cell(row=current_row, column=1, value=row_data[0]).font = Font(bold=True)
                    ws.cell(row=current_row, column=2, value=row_data[1])
                    current_row += 1

                current_row += 1  # Пустая строка после информации о соревновании

                # Заголовки таблицы участников
                headers = [
                    "ID участника", "ФИО", "Никнейм", "Регион", 
                    "Роль", "Рейтинг", "Место", "Результат", "Тип участия"
                ]
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=header)
                    cell.font = header_font
                    cell.border = border

                current_row += 1

                # Участники (индивидуальные)
                participants = CompetitionParticipant.objects.filter(competition=comp).select_related(
                    'participant', 'participant__user', 'participant__region', 'participant__role'
                )
                
                for participant in participants:
                    user_info = participant.participant
                    result = CompetitionResult.objects.filter(
                        competition=comp,
                        participant=user_info
                    ).first()
                    
                    participant_data = [
                        user_info.user.id,
                        f"{user_info.surname} {user_info.name} {user_info.patronymic or ''}".strip(),
                        user_info.user.nickName,
                        user_info.region.name,
                        user_info.role.name,
                        user_info.rating,
                        result.place if result else "-",
                        participant.result or "-",
                        "Индивидуальный"
                    ]
                    
                    for col_num, value in enumerate(participant_data, 1):
                        cell = ws.cell(row=current_row, column=col_num, value=value)
                        cell.border = border
                    current_row += 1

                # Команды
                teams = Team.objects.filter(competition=comp).prefetch_related(
                    'members', 'members__user', 'members__region', 'members__role'
                )
                
                for team in teams:
                    # Заголовок команды
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                    cell = ws.cell(row=current_row, column=1, value=f"Команда: {team.name} (Капитан: {team.captain.user.nickName if team.captain else 'не указан'})")
                    cell.font = Font(bold=True, italic=True)
                    current_row += 1

                    for member in team.members.all():
                        result = CompetitionResult.objects.filter(
                            competition=comp,
                            participant=member
                        ).first()
                        
                        member_data = [
                            member.user.id,
                            f"{member.surname} {member.name} {member.patronymic or ''}".strip(),
                            member.user.nickName,
                            member.region.name,
                            member.role.name,
                            member.rating,
                            result.place if result else "-",
                            "-",  # Для командных результатов
                            "Командный"
                        ]
                        
                        for col_num, value in enumerate(member_data, 1):
                            cell = ws.cell(row=current_row, column=col_num, value=value)
                            cell.border = border
                        current_row += 1

                current_row += 3  # Отступ перед следующим соревнованием

            # Настраиваем ширину столбцов (обход объединенных ячеек)
            for col in range(1, 10):  # У нас 9 столбцов
                max_length = 0
                column_letter = get_column_letter(col)
                
                # Проверяем только необъединенные ячейки
                for row in range(1, current_row + 1):
                    try:
                        cell = ws.cell(row=row, column=col)
                        if cell.value and not isinstance(cell, ws.merged_cells):
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем файл
            wb.save(output)
            output.seek(0)

            # Настраиваем HTTP ответ
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            filename = 'structured_competitions_participants.xlsx' if not competition_id else f'structured_competition_{competition_id}_participants.xlsx'
            response['Content-Disposition'] = f'attachment; filename={filename}'
            
            return response
            
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )